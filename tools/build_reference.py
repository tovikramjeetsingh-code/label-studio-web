#!/usr/bin/env python3
"""Build assets/reference.enc.js — the ENCRYPTED SKU reference — from every
Seller Listings Report CSV/XLSX in private_source/.

The catalog is encrypted with the shared team password (AES-256-GCM, key derived
via PBKDF2-HMAC-SHA256) so the committed file is safe to publish: without the
password it is unreadable ciphertext. The browser decrypts it after the user
enters the password once (assets/parse.js decryptReference).

Usage:
    python3 tools/build_reference.py --password "the team password"
    # or set REF_PASSWORD in the environment

Re-run whenever listings change or you rotate the password, then commit + push.
private_source/ is git-ignored — raw catalogs are never committed.
"""
import argparse, base64, csv, glob, json, os, datetime, re, sys

try:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes
except ImportError:
    sys.exit("Missing dependency: pip install cryptography")

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "..", "private_source")
OUT = os.path.join(HERE, "..", "assets", "reference.enc.js")
ITERS = 200_000

COLS = {"brand": "b", "article type": "a", "style name": "sn", "style id": "si",
        "size": "sz", "seller sku code": "ss", "sku code": "sk", "mrp": "m",
        "van": "v"}
ALIASES = {
    "brand": ["brand"], "article type": ["article type"], "style name": ["style name"],
    "style id": ["style id"], "size": ["size"],
    "seller sku code": ["seller sku code", "seller sku"],
    "sku code": ["sku code", "sku_code"], "mrp": ["mrp"],
    "van": ["van", "vendor article no", "vendor article number"],
}


def clean(v):
    s = ("" if v is None else str(v)).strip()
    if s.lower() == "nan":
        return ""
    if re.fullmatch(r"\d+\.0+", s):   # 1999.0 / 1999.0000 -> 1999
        s = s.split(".")[0]
    return s


def read_rows(path):
    if path.lower().endswith((".xlsx", ".xls", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Missing dependency for Excel: pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c) if c is not None else "" for c in rows[0]]
        return hdr, [dict(zip(hdr, r)) for r in rows[1:]]
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        return reader.fieldnames, list(reader)


def build_master():
    files = sorted(glob.glob(os.path.join(SRC, "*.csv")) + glob.glob(os.path.join(SRC, "*.xlsx")))
    if not files:
        sys.exit(f"No listing files in {os.path.abspath(SRC)} — move the Seller Listings Reports there first.")
    records, by_seller, by_sku, by_skuid, sources = [], {}, {}, {}, []
    seen = set()

    # Some listings carry the SKU code in the seller-sku field, so the two are
    # identical. When the same SKU is also listed properly (usually on the other
    # account) that bad row must lose, or a lookup by SKU code hands back a
    # seller SKU that matches nothing in OMS or an STN — which silently breaks
    # reconcile and the scan tab. Work out which SKUs have a proper row first.
    proper = set()
    for path in files:
        _hdr, _rows = read_rows(path)
        lut = {(c or "").strip().lower().rstrip(":"): c for c in _hdr}
        sc = next((lut[a] for a in ALIASES["sku code"] if a in lut), None)
        ss = next((lut[a] for a in ALIASES["seller sku code"] if a in lut), None)
        if not sc or not ss:
            continue
        for r in _rows:
            a, b = clean(r.get(sc)), clean(r.get(ss))
            if a and b and a.lower() != b.lower():
                proper.add(a.lower())
    dropped_dupes = 0
    for path in files:
        hdr, rows = read_rows(path)
        lut = {(c or "").strip().lower().rstrip(":"): c for c in hdr}
        skuid_col = lut.get("sku id")
        n = 0
        for r in rows:
            rec = {}
            for canon, short in COLS.items():
                col = next((lut[a] for a in ALIASES[canon] if a in lut), None)
                rec[short] = clean(r.get(col)) if col else ""
            if not rec["ss"] and not rec["sk"]:
                continue
            # The same SKU is often listed on both seller accounts with identical
            # label fields — keep one record so the finder doesn't show (and
            # print) every size twice.
            # seller sku == sku code, and a proper listing exists for it -> skip
            if (rec["sk"] and rec["ss"].lower() == rec["sk"].lower()
                    and rec["sk"].lower() in proper):
                dropped_dupes += 1
                continue
            dedupe_key = (rec["ss"].lower(), rec["sk"].lower())
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            idx = len(records); records.append(rec); n += 1
            if rec["ss"]:
                by_seller[rec["ss"].lower()] = idx
            if rec["sk"]:
                by_sku[rec["sk"].lower()] = idx
            sid = clean(r.get(skuid_col)).lower() if skuid_col else ""
            if sid:
                by_skuid[sid] = idx
        sources.append(f"{os.path.basename(path)} ({n})")
    if dropped_dupes:
        print(f"  dropped {dropped_dupes} rows whose seller sku code == sku code "
              f"(a proper listing exists for the same SKU)")
    return {
        "meta": {"count": len(records), "built": datetime.date.today().isoformat(), "sources": sources},
        "records": records, "bySeller": by_seller, "bySku": by_sku, "bySkuId": by_skuid,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--password", default=os.environ.get("REF_PASSWORD", ""))
    args = ap.parse_args()
    if not args.password:
        sys.exit("Provide the team password via --password or REF_PASSWORD env var.")

    master = build_master()
    plaintext = json.dumps(master, ensure_ascii=False, separators=(",", ":")).encode("utf-8")

    salt = os.urandom(16)
    iv = os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=ITERS)
    key = kdf.derive(args.password.encode("utf-8"))
    ct = AESGCM(key).encrypt(iv, plaintext, None)   # ciphertext || tag

    enc = {
        "v": 1, "iters": ITERS,
        "salt": base64.b64encode(salt).decode(),
        "iv": base64.b64encode(iv).decode(),
        "ct": base64.b64encode(ct).decode(),
    }
    with open(OUT, "w", encoding="utf-8") as out:
        out.write("// Encrypted SKU reference — generated by tools/build_reference.py.\n")
        out.write("// Public ciphertext; unreadable without the team password.\n")
        out.write("window.LABEL_ENC = ")
        json.dump(enc, out, separators=(",", ":"))
        out.write(";\n")
    print(f"Wrote {OUT}")
    print(f"  {master['meta']['count']} SKUs from {', '.join(master['meta']['sources'])}")
    print(f"  encrypted {len(plaintext)} bytes -> {len(enc['ct'])} b64 chars (AES-256-GCM, PBKDF2 {ITERS})")


if __name__ == "__main__":
    main()
